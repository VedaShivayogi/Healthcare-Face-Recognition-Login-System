"""
KLIKE v4 – Export engine.
Generates Excel (.xlsx) and PDF reports from logs and patient data.
"""

import os
from datetime import datetime

ROOT       = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
REPORT_DIR = os.path.join(ROOT, "reports")
os.makedirs(REPORT_DIR, exist_ok=True)

def _stamp():
    return datetime.now().strftime("%Y%m%d_%H%M%S")


# ── EXCEL ─────────────────────────────────────────────────────────────────────

def export_logs_excel(logs):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "openpyxl not installed. Run: pip install openpyxl"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Access Logs"

    # Style helpers
    teal_fill  = PatternFill("solid", fgColor="00D4C8")
    head_font  = Font(bold=True, color="0A1628", size=10)
    bold_font  = Font(bold=True)

    headers = ["#", "Timestamp", "Name", "Role", "Status", "Method"]
    widths  = [5, 22, 20, 12, 12, 12]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = head_font
        cell.fill      = teal_fill
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = w

    for ri, log in enumerate(logs, 2):
        row_data = [ri-1, log.get("timestamp",""), log.get("name",""),
                    log.get("role",""), log.get("status",""), log.get("method","")]
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(horizontal="left")
            if log.get("status") == "granted":
                cell.font = Font(color="00897B")
            elif log.get("status") == "denied":
                cell.font = Font(color="FF4C6A")

    path = os.path.join(REPORT_DIR, f"access_log_{_stamp()}.xlsx")
    wb.save(path)
    return path, None


def export_patients_excel(patients):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None, "openpyxl not installed"

    wb  = openpyxl.Workbook()
    ws  = wb.active
    ws.title = "Patients"
    teal_fill = PatternFill("solid", fgColor="00D4C8")
    head_font = Font(bold=True, color="0A1628")

    headers = ["Name", "DOB", "Blood Type", "Doctor", "Conditions", "Appointments", "Registered"]
    widths  = [20, 14, 12, 18, 25, 14, 14]
    for ci, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = head_font; cell.fill = teal_fill
        ws.column_dimensions[cell.column_letter].width = w

    for ri, (name, info) in enumerate(patients.items(), 2):
        row_data = [
            name, info.get("dob",""), info.get("blood_type",""),
            info.get("doctor",""), info.get("conditions",""),
            len(info.get("appointments",[])), info.get("registered","")
        ]
        for ci, val in enumerate(row_data, 1):
            ws.cell(row=ri, column=ci, value=val)

    path = os.path.join(REPORT_DIR, f"patients_{_stamp()}.xlsx")
    wb.save(path)
    return path, None


# ── PDF ───────────────────────────────────────────────────────────────────────

def export_logs_pdf(logs):
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.units import cm
    except ImportError:
        return None, "reportlab not installed. Run: pip install reportlab"

    path = os.path.join(REPORT_DIR, f"access_log_{_stamp()}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=landscape(A4),
                              leftMargin=1.5*cm, rightMargin=1.5*cm,
                              topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    story  = []
    story.append(Paragraph("<b>KLIKE Healthcare – Access Log Report</b>",
                            styles["Title"]))
    story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                            styles["Normal"]))
    story.append(Spacer(1, 0.4*cm))

    data = [["#", "Timestamp", "Name", "Role", "Status", "Method"]]
    for i, log in enumerate(logs, 1):
        data.append([i, log.get("timestamp",""), log.get("name",""),
                     log.get("role",""), log.get("status",""), log.get("method","")])

    teal  = colors.HexColor("#00D4C8")
    green = colors.HexColor("#00897B")
    red   = colors.HexColor("#FF4C6A")
    navy  = colors.HexColor("#0A1628")

    ts = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), teal),
        ("TEXTCOLOR",   (0,0), (-1,0), navy),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#EEF4FB")]),
        ("GRID",        (0,0), (-1,-1), 0.4, colors.HexColor("#C5D8F0")),
        ("LEFTPADDING", (0,0), (-1,-1), 6),
        ("RIGHTPADDING",(0,0), (-1,-1), 6),
    ])
    # Colour status column
    for ri, log in enumerate(logs, 1):
        col = green if log.get("status") == "granted" else red
        ts.add("TEXTCOLOR", (4, ri), (4, ri), col)
        ts.add("FONTNAME",  (4, ri), (4, ri), "Helvetica-Bold")

    col_widths = [1*cm, 4.5*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm]
    tbl = Table(data, colWidths=col_widths, repeatRows=1)
    tbl.setStyle(ts)
    story.append(tbl)
    doc.build(story)
    return path, None


def export_patient_pdf(name, info):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
        from reportlab.lib.units import cm
    except ImportError:
        return None, "reportlab not installed"

    path = os.path.join(REPORT_DIR, f"patient_{name.replace(' ','_')}_{_stamp()}.pdf")
    doc  = SimpleDocTemplate(path, pagesize=A4,
                              leftMargin=2*cm, rightMargin=2*cm,
                              topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    teal   = colors.HexColor("#00D4C8")
    navy   = colors.HexColor("#0A1628")
    story  = []

    story.append(Paragraph("<b>KLIKE Healthcare</b>", styles["Title"]))
    story.append(Paragraph("Patient Access Report", styles["Heading2"]))
    story.append(HRFlowable(width="100%", color=teal, thickness=2))
    story.append(Spacer(1, 0.3*cm))

    details = [
        ["Patient Name:", name],
        ["Date of Birth:", info.get("dob","—")],
        ["Blood Type:",   info.get("blood_type","—")],
        ["Conditions:",   info.get("conditions","—")],
        ["Primary Doctor:", info.get("doctor","—")],
        ["Registered:",   info.get("registered","—")],
    ]
    tbl = Table(details, colWidths=[4.5*cm, 11*cm])
    tbl.setStyle(TableStyle([
        ("FONTNAME",    (0,0), (0,-1), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 9),
        ("TEXTCOLOR",   (0,0), (0,-1), teal),
        ("ROWBACKGROUNDS",(0,0),(-1,-1),[colors.white, colors.HexColor("#EEF4FB")]),
        ("LEFTPADDING", (0,0),(-1,-1), 6),
        ("TOPPADDING",  (0,0),(-1,-1), 4),
        ("BOTTOMPADDING",(0,0),(-1,-1), 4),
    ]))
    story.append(tbl)
    story.append(Spacer(1, 0.4*cm))

    appts = info.get("appointments", [])
    if appts:
        story.append(Paragraph("<b>Appointments</b>", styles["Heading3"]))
        adata = [["Date","Time","Department","Doctor","Status"]]
        for a in appts:
            adata.append([a.get("date",""), a.get("time",""),
                          a.get("dept",""), a.get("doctor",""), a.get("status","")])
        atbl = Table(adata, colWidths=[2.5*cm,2*cm,4*cm,4*cm,3*cm])
        atbl.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0), teal),
            ("TEXTCOLOR", (0,0),(-1,0), navy),
            ("FONTNAME",  (0,0),(-1,0), "Helvetica-Bold"),
            ("FONTSIZE",  (0,0),(-1,-1), 8),
            ("GRID",      (0,0),(-1,-1), 0.4, colors.HexColor("#C5D8F0")),
        ]))
        story.append(atbl)

    notes = info.get("notes",[])
    if notes:
        story.append(Spacer(1, 0.3*cm))
        story.append(Paragraph("<b>Clinical Notes</b>", styles["Heading3"]))
        for n in notes:
            story.append(Paragraph(f"<i>{n.get('date','')} – {n.get('author','')}</i>: {n.get('text','')}",
                                   styles["Normal"]))
            story.append(Spacer(1, 0.1*cm))

    doc.build(story)
    return path, None
